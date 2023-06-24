import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options

# 打开 Excel 文件
workbook_wallet = openpyxl.load_workbook('./wallet.xlsx')

# 获取默认的工作表
worksheet = workbook_wallet.active

# 获取第二行数据
row = 2  # 第二行的索引为2

Address = worksheet.cell(row=row, column=1).value
Seed_Phrase = worksheet.cell(row=row, column=2).value
PrivateKey = worksheet.cell(row=row, column=3).value

print(Address, Seed_Phrase, PrivateKey)

# 创建 ChromeOptions 对象
options = Options()

# 添加插件路径
plugin_path = './metamask.crx'
options.add_extension(plugin_path)

# 设置 Chrome WebDriver 的路径
chromedriver_path = '/path/to/chromedriver'

# 创建 Chrome WebDriver，并传递 options 参数
driver = webdriver.Chrome(options=options)
driver.set_window_size(1280, 900)
##
# 等待出现2个窗口，等待时间为10秒
WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
# 通过窗口句柄中，检测窗口1标题是否为metamask
metamask_handle = driver.window_handles[1]
# 切换metamask窗口
driver.switch_to.window(metamask_handle)
# 判断窗口url中是否含有home文本，等待时间为5秒
WebDriverWait(driver, 5).until(EC.url_contains('home'))

# 等待账户输入元素
el_1 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                         '//*[@id="onboarding__terms-checkbox"]'))
el_1.click()
#
el_2 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                         '//*[@id="app-content"]/div/div[2]/div/div/div/ul/li[3]/button'))
el_2.click()
ea_1 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                         '//*[@id="app-content"]/div/div[2]/div/div/div/div/button[1]'))
ea_1.click()
input()

# 拆分助记词为单词列表
seed_phrase_words = Seed_Phrase.split()

# 定位助记词输入框并填入助记词
for index, word in enumerate(seed_phrase_words):
    input_xpath_1 = f'//*[@id="import-srp__srp-word-{index}"]'
    input_element = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_1))
    input_element.send_keys(word)

# 定位并点击指定的元素
el_3 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[4]/div/button'
button_element = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, el_3))
button_element.click()

# 输入密码导入
mm = 'aa123456'
input_xpath_2 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/div[1]/label/input'
el_4 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_2))
el_4.send_keys(mm)
# 确认密码
input_xpath_3 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/div[2]/label/input'
el_5 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_3))
el_5.send_keys(mm)
# 打钩
input_xpath_4 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/div[3]/label/input'
el_6 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_4))
el_6.click()

# 导入我的钱包
input_xpath_5 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/button'
el_7 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_5))
el_7.click()

##

input_xpath_6 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/button'
el_8 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_6))
el_8.click()
##

input_xpath_7 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/button'
el_9 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_7))
el_9.click()

##

input_xpath_8 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/button'
el_10 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_8))
el_10.click()

###清除弹窗
input_xpath_9 = '//*[@id="popover-content"]/div/div/section/div[2]/div/button/span'
el_9 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_99))
el_9.click()

#####
time.sleep(100000)
###########BNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNBBNB
'''
    #切换网络BNB
input_xpath_9 = '//*[@id="app-content"]/div/div[1]/div/div[2]/div/div/span'
el_11 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_9))
el_11.click()

    #添加网络
input_xpath_10 = '//*[@id="app-content"]/div/div[2]/div/div[3]/button'
el_12 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_10))
el_12.click()

#添加BNB
input_xpath_11 = '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div[4]/div[2]/button'
el_13 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_11))
el_13.click()
#
input_xpath_12 = '//*[@id="popover-content"]/div/div/section/div/div/div[2]/div/button[2]'
el_14 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_12))
el_14.click()
#
input_xpath_13 = '//*[@id="popover-content"]/div/div/section/div/div/button[2]'
el_15 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_13))
el_15.click()
#切到BNB网络
input_xpath_9 = '//*[@id="app-content"]/div/div[1]/div/div[2]/div/div/span'
el_11 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_9))
el_11.click()
#
input_xpath_14 = '//*[@id="app-content"]/div/div[2]/div/div[2]/li[2]/span'
el_16 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_14))
el_16.click()
#
input_xpath_15 = '//*[@id="popover-content"]/div/div/section/div[3]/button'
el_17 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_15))
el_17.click()
'''
################################################################################

import openpyxl

# 读取 xlsx 文件
workbook_tw = openpyxl.load_workbook('./tw_accounts.xlsx')
# 获取其中所有的工作表
print(workbook_tw.sheetnames)
# 打开第一个工作表
table = workbook_tw.active

# 创建新的 Excel 文件，用于记录登录失败的账户
fail_workbook = openpyxl.Workbook()
fail_sheet = fail_workbook.active
fail_row = 1

# 遍历数据行
for row in table.iter_rows(min_row=2):
    row_values = [cell.value for cell in row]
    
    username = row_values[0]
    password = row_values[1]
    token = row_values[2]
    
    print(username, password, token)
    
    # 打开网站url
    driver.get('https://twitter.com/i/flow/login')
    
    # 等待账户输入元素
    el_1 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                             '//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div/div/div[5]/label/div/div[2]/div/input'))
    # 输入
    el_1.send_keys(username)
    # 定位下一步
    el_2 = driver.find_element(By.XPATH,
                               '//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div/div/div[6]/div')
    el_2.click()
    # 等待密码框
    el_3 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                             '//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div[1]/div/div/div[3]/div/label/div/div[2]/div[1]/input'))
    el_3.send_keys(password)
    # 登录
    el_4 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                             '//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div[2]/div/div[1]/div/div/div/div'))
    try:
        el_4.click()
        # 输入邮箱
        el_4 = WebDriverWait(driver, timeout=5).until(lambda d: d.find_element(By.XPATH,
                                                                               '//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div[1]/div/div[2]/label/div/div[2]/div/input'))
        el_4.send_keys(email)
        # 下一步
        el_5 = driver.find_element(By.XPATH,
                                   '//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div[2]/div/div/div/div/div')
        el_5.click()
        WebDriverWait(driver, timeout=10).until(EC.url_contains("https://twitter.com/home"))
        print("登录成功")
    
    except TimeoutException:
        
        print("未找到邮箱输入框，登录成功")
    
    except Exception as e:
        
        print("登录失败")
        fail_sheet.cell(fail_row, 1, value=row_values[0])
        fail_sheet.cell(fail_row, 2, value=row_values[1])
        fail_sheet.cell(fail_row, 3, value=row_values[2])
        fail_row += 1
        continue  # 跳过当前账号登录的循环，尝试下一个账号的登录
        
        # 在副本中删除该行
    table.delete_rows(i)
    workbook_tw.save('./tw_accounts.xlsx')
    
    i -= 1

# 保存登录失败的账户信息到新 Excel 文件中
fail_workbook.save('./tw_login_fail.xlsx')

#######Discord登录######################

# 打开原始 Excel 文件
workbook_dc = openpyxl.load_workbook('./dc_accounts.xlsx')
# 创建一个可写的副本
workbook_dc_w = openpyxl.load_workbook('./dc_accounts.xlsx')
# 获取其中所有的工作表
print(workbook_dc.sheetnames)
# 打开第一个工作表
table = workbook_dc['Sheet1']
# 获取第一个工作表的行数
num_rows = table.max_row
# 创建新的 Excel 文件，用于记录登录失败的账户
fail_workbook = openpyxl.Workbook()
fail_sheet = fail_workbook.active
fail_sheet.title = 'login_fail'
fail_row = 1

for i in range(1, table.max_row):
    dc_email = table.cell(row=i, column=1).value
    dc_password = table.cell(row=i, column=2).value
    dc_token = table.cell(row=i, column=3).value
    print(dc_email, dc_password, dc_token)
    
    driver.execute_script("window.open('about:blank', 'new_window')")
    driver.switch_to.window("new_window")
    driver.get('https://discord.com/login')
    token = dc_token
    # 在当前页面中注入 JavaScript 代码
    driver.execute_script(f"""
 (function () {{
      const token = "{token}";
      window.localStorage = document.body.appendChild(document.createElement("iframe")).contentWindow.localStorage;
      window.setInterval(() => document.defaultView.localStorage.token = `"{token}"`);
      window.location.reload();
 }})();
 """)
    try:
        WebDriverWait(driver, timeout=10).until(EC.url_contains("https://discord.com/channels/@me"))
    except TimeoutException:
        print("登录失败")
        fail_sheet.cell(row=fail_row, column=1, value=table.cell(row=i, column=1).value)
        fail_sheet.cell(row=fail_row, column=2, value=table.cell(row=i, column=2).value)
        fail_sheet.cell(row=fail_row, column=3, value=table.cell(row=i, column=3).value)
        fail_row += 1
        continue  # 跳过当前账号登录的循环，尝试下一个账号的登录
    else:
        print("登录成功")
    finally:
        # 关闭文件
        workbook_dc.save('./dc_accounts.xlsx')
    
    # 删除已登录的账号
    sheet_name = 'Sheet1'
    if sheet_name in workbook_dc.sheetnames:
        sheet_to_remove = workbook_dc[sheet_name]
        for j in range(sheet_to_remove.max_row, 0, -1):
            if sheet_to_remove.cell(row=j, column=3).value == '':
                continue
            elif sheet_to_remove.cell(row=j, column=3).value == dc_token:
                sheet_to_remove.delete_rows(j, 1)
                break
        workbook_dc.save('./dc_accounts.xlsx')

input()
